import pandas as pd
from openpyxl.styles import Alignment, Font

from .named_tuple import isinstance_namedtuple


def loadcases2excel(xl_filename, dfObj, ws_header_lines=4):
    """
    """
    def make_transposed_df(_data):
        """Pandas work-around"""
        try:
            df = pd.DataFrame(data=_data)   
        except:     # ValueError:
            df = pd.DataFrame(data=_data, index=[0]) # ValueError: If using all scalar values, you must pass an index    
        return df.T        
    with pd.ExcelWriter(xl_filename, engine="openpyxl", mode='w') as writer:
        for wsidx, spec in enumerate(dfObj):
            dataObj, sheetName, *rowStrings = spec
            #print(wsidx, rowStrings)
            if isinstance(dataObj, dict):
                _data = dataObj
                df = make_transposed_df(_data)
            elif isinstance_namedtuple(dataObj):
                #_data = dataObj._asdict()
                _data = {type(dataObj).__name__: ""}
                _data.update(dataObj._asdict()) 
                df = make_transposed_df(_data)
            elif isinstance(dataObj, (list, tuple)) and isinstance_namedtuple(dataObj[0]):
                # _data = {}
                # for nt in dataObj:
                #     _data[type(nt).__name__] = ""
                #     _data.update(nt._asdict())
                df = pd.DataFrame()
                for rnt in dataObj:
                    _data = {type(rnt).__name__: ""}
                    _data.update(rnt._asdict()) 
                    _df = make_transposed_df(_data)
                    df = pd.concat([df, _df])
            else:
                continue
            #with pd.ExcelWriter("atestpipeline.xlsx", engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheetName)
            ws = writer.book.worksheets[-1]
            ws.insert_rows(0, ws_header_lines)
            for jj, txt in enumerate(rowStrings):
                if jj >= ws_header_lines:
                    break
                ws.cell(row=jj+1, column=1).value = txt
            for jj, _cell in enumerate(ws['A']):
                _cell.alignment = Alignment(horizontal='left')
                if 1 <= jj <= ws_header_lines:
                    continue
                _cell.font = Font(bold=True)
            #ws['A1'].font = Font(bold=True)


#local_buckling_propagation_all.__code__.co_varnames.index("kwargs")